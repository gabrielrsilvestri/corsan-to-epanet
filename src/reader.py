"""
src/reader.py
─────────────────────────────────────────────────────────────────────────────
Lê CSV de coordenadas + XLSX de dimensionamento CORSAN e retorna um objeto
Network pronto para serialização ou geração de INP.

Uso:
    python -m src.reader config/gravatai_p96.yaml
─────────────────────────────────────────────────────────────────────────────
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import warnings
from pathlib import Path
from typing import Any

import openpyxl
import yaml

from src.models import Network, NetworkMetadata, Node, Pipe


# Mapeamento de colunas da planilha (1-based, conforme CLAUDE.md)
COL_MAP: dict[str, int] = {
    "montante":           1,
    "jusante":            2,
    "ni_trecho":          3,
    "diametro_mm":        4,
    "comprimento_m":      5,
    "n_contrib_montante": 6,
    "n_contrib_jusante":  7,
    "vazao_ls":           8,
    "vazao_m3s":          9,
    "velocidade_ms":     10,
    "perda_unit_mKm":    11,
    "perda_total_m":     12,
    "cota_montante":     13,
    "cota_jusante":      14,
    "desnivel_terreno":  15,
    "pressao_disp_mont": 16,
    "pressao_disp_jus":  17,
    "pressao_est_mont":  18,
    "pressao_est_jus":   19,
    "nivel_piezo_mont":  20,
    "nivel_piezo_jus":   21,
}


def _normalize_id(s: str) -> str:
    """Normaliza ID de nó: IDs puramente numéricos perdem zeros à esquerda.
    Ex: '01' → '1', '007' → '7', 'PT' → 'PT'.
    """
    try:
        return str(int(s))
    except ValueError:
        return s


def _to_float(value: Any) -> float | None:
    """Converte valor de célula Excel para float, tratando vírgula decimal."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_config(config_path: Path) -> dict:
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _read_coordinates(csv_path: Path) -> dict[str, tuple[float, float]]:
    """Lê CSV de coordenadas e retorna mapa node_id → (easting, northing)."""
    coords: dict[str, tuple[float, float]] = {}
    for encoding in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            with open(csv_path, encoding=encoding, newline="") as f:
                f.read(2048)
            break
        except UnicodeDecodeError:
            continue
    else:
        encoding = "latin-1"

    with open(csv_path, encoding=encoding, newline="") as f:
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        for row in reader:
            row = {k.strip(): v.strip() for k, v in row.items() if k}
            name = row.get("Name") or row.get("name")
            raw_e = row.get("Easting") or row.get("easting")
            raw_n = row.get("Northing") or row.get("northing")
            if not (name and raw_e and raw_n):
                continue
            try:
                coords[_normalize_id(str(name))] = (
                    float(raw_e.replace(",", ".")),
                    float(raw_n.replace(",", ".")),
                )
            except ValueError:
                warnings.warn(
                    f"Coordenada inválida para nó '{name}': "
                    f"E={raw_e}, N={raw_n}"
                )
    return coords


def _read_meta(ws, meta_cells: dict) -> dict[str, float | None]:
    return {
        field: _to_float(ws.cell(row=row, column=col).value)
        for field, (row, col) in meta_cells.items()
    }


def _read_xlsx_rows(ws, data_start_row: int, col_offset: int = 0) -> list[dict]:
    """Lê linhas de trechos até a primeira célula vazia na coluna de montante."""
    rows: list[dict] = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        if not ws.cell(row=row_idx, column=COL_MAP["montante"] + col_offset).value:
            break
        rows.append(
            {field: ws.cell(row=row_idx, column=col + col_offset).value
             for field, col in COL_MAP.items()}
        )
    return rows


def build_network(config: dict, base_dir: Path | None = None) -> Network:
    """
    Constrói e retorna um Network a partir do CSV e XLSX indicados no config.

    base_dir: diretório de referência para resolver paths relativos no config.
              Padrão: diretório de trabalho atual.
    """
    if base_dir is None:
        base_dir = Path.cwd()

    csv_path       = base_dir / config["csv_path"]
    xlsx_path      = base_dir / config["xlsx_path"]
    sheet_name     = config.get("sheet_name")
    data_start_row = int(config.get("data_start_row", 13))
    col_offset     = int(config.get("col_offset", 0))
    reservoir_id   = str(config.get("reservoir_id", "PT"))
    meta_cells     = config.get("meta_cells", {})

    # 1. Coordenadas ──────────────────────────────────────────────────────────
    coords = _read_coordinates(csv_path)

    # 2. Planilha XLSX ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        raise ValueError(f"Planilha não encontrada em '{xlsx_path}'")

    sheet_title  = ws.title or ""
    meta_values  = _read_meta(ws, meta_cells)
    rows         = _read_xlsx_rows(ws, data_start_row, col_offset)
    wb.close()

    # 3. Metadados ─────────────────────────────────────────────────────────────
    metadata = NetworkMetadata(
        projeto_nome=config.get("projeto_nome", ""),
        fonte_coordenadas=str(csv_path),
        fonte_dimensionamento=str(xlsx_path),
        aba_dimensionamento=sheet_title,
        N_economias=meta_values.get("N_economias"),
        C_HW=meta_values.get("C_HW"),
        Q_unit_ls=meta_values.get("Q_unit_ls"),
        pressao_PT_mca=meta_values.get("pressao_PT_mca"),
        reservoir_id=reservoir_id,
    )

    Q_unit = metadata.Q_unit_ls or 0.0
    C_HW   = metadata.C_HW or 150.0

    # N_economias: prevalece o maior n_contrib_jusante encontrado nos trechos
    n_contrib_values = [
        _to_float(row["n_contrib_jusante"])
        for row in rows
        if _to_float(row["n_contrib_jusante"]) is not None
    ]
    n_economias_dim = max(n_contrib_values) if n_contrib_values else None
    if n_economias_dim is not None and metadata.N_economias != n_economias_dim:
        print(
            f"AVISO: N_economias do cabeçalho ({metadata.N_economias}) difere do "
            f"dimensionamento ({n_economias_dim}). Adotando {n_economias_dim}.",
            file=sys.stderr,
        )
    if n_economias_dim is not None:
        metadata.N_economias = n_economias_dim

    # 4. Trechos (Pipes) ───────────────────────────────────────────────────────
    pipes: list[Pipe] = []
    pipe_id_count: dict[str, int] = {}

    for row in rows:
        node1_id = _normalize_id(str(row["montante"]).strip())
        node2_id = _normalize_id(str(row["jusante"]).strip())

        base_id  = f"P_{node1_id}_{node2_id}"
        count    = pipe_id_count.get(base_id, 0) + 1
        pipe_id_count[base_id] = count
        pipe_id  = base_id if count == 1 else f"{base_id}_{count}"

        pipes.append(Pipe(
            id=pipe_id,
            node1=node1_id,
            node2=node2_id,
            comprimento_m=_to_float(row["comprimento_m"]),
            diametro_mm=_to_float(row["diametro_mm"]),
            rugosidade_HW=C_HW,
            ni_trecho=_to_float(row["ni_trecho"]),
            n_contrib_montante=_to_float(row["n_contrib_montante"]),
            n_contrib_jusante=_to_float(row["n_contrib_jusante"]),
            vazao_ls=_to_float(row["vazao_ls"]),
            vazao_m3s=_to_float(row["vazao_m3s"]),
            velocidade_ms=_to_float(row["velocidade_ms"]),
            perda_unit_mKm=_to_float(row["perda_unit_mKm"]),
            perda_total_m=_to_float(row["perda_total_m"]),
        ))

    # 5. Nós ───────────────────────────────────────────────────────────────────
    # Acumula propriedades de cada nó na ordem de primeira aparição
    ordered_ids:    list[str]              = []
    elevation:      dict[str, float|None] = {}
    n_contrib_map:  dict[str, float|None] = {}
    pressao_disp:   dict[str, float|None] = {}
    pressao_est:    dict[str, float|None] = {}
    nivel_piezo:    dict[str, float|None] = {}
    reservoir_head: dict[str, float|None] = {}

    def _register(node_id: str) -> None:
        if node_id not in elevation:
            ordered_ids.append(node_id)
            elevation[node_id]     = None
            n_contrib_map[node_id] = None
            pressao_disp[node_id]  = None
            pressao_est[node_id]   = None
            nivel_piezo[node_id]   = None

    def _set_first(d: dict, key: str, value: float | None) -> None:
        if d[key] is None:
            d[key] = value

    for row, pipe in zip(rows, pipes):
        m, j = pipe.node1, pipe.node2
        _register(m)
        _register(j)

        cota_m = _to_float(row["cota_montante"])
        cota_j = _to_float(row["cota_jusante"])

        # Elevação: primeira ocorrência prevalece; divergência > 0.01 m → aviso
        if elevation[m] is None:
            elevation[m] = cota_m
        elif cota_m is not None and abs((elevation[m] or 0.0) - cota_m) > 0.01:
            warnings.warn(
                f"Nó '{m}': cota divergente "
                f"(anterior={elevation[m]:.3f}, atual={cota_m:.3f})"
            )

        if elevation[j] is None:
            elevation[j] = cota_j
        elif cota_j is not None and abs((elevation[j] or 0.0) - cota_j) > 0.01:
            warnings.warn(
                f"Nó '{j}': cota divergente "
                f"(anterior={elevation[j]:.3f}, atual={cota_j:.3f})"
            )

        # n_contrib do nó vem de n_contrib_jusante quando ele é jusante
        nc_j = _to_float(row["n_contrib_jusante"])
        if n_contrib_map[j] is None and nc_j is not None:
            n_contrib_map[j] = nc_j

        # Pressão e nível piezométrico (primeira ocorrência)
        _set_first(pressao_disp, m, _to_float(row["pressao_disp_mont"]))
        _set_first(pressao_est,  m, _to_float(row["pressao_est_mont"]))
        _set_first(nivel_piezo,  m, _to_float(row["nivel_piezo_mont"]))
        _set_first(pressao_disp, j, _to_float(row["pressao_disp_jus"]))
        _set_first(pressao_est,  j, _to_float(row["pressao_est_jus"]))
        _set_first(nivel_piezo,  j, _to_float(row["nivel_piezo_jus"]))

        # total_head do reservatório = nivel_piezo_mont do primeiro trecho
        # onde ele aparece como montante; fallback: cota + pressao_disp
        if m == reservoir_id and reservoir_head.get(reservoir_id) is None:
            piezo = _to_float(row["nivel_piezo_mont"])
            if piezo is None:
                cota = _to_float(row["cota_montante"])
                p    = _to_float(row["pressao_disp_mont"])
                piezo = (cota + p) if (cota is not None and p is not None) else None
            reservoir_head[reservoir_id] = piezo

    # 5b. Demanda local por balanço de fluxo: Q_entrando - sum(Q_saindo)
    arriving_q: dict[str, float] = {}
    leaving_q:  dict[str, float] = {}
    for pipe in pipes:
        leaving_q[pipe.node1]  = leaving_q.get(pipe.node1, 0.0) + (pipe.vazao_ls or 0.0)
        arriving_q[pipe.node2] = arriving_q.get(pipe.node2, 0.0) + (pipe.vazao_ls or 0.0)

    nodes: list[Node] = []
    for node_id in ordered_ids:
        is_res = node_id == reservoir_id
        nc     = n_contrib_map[node_id]
        if is_res:
            demand = 0.0
        else:
            demand = arriving_q.get(node_id, 0.0) - leaving_q.get(node_id, 0.0)
            demand = max(0.0, round(demand, 6))
        coord  = coords.get(node_id)

        nodes.append(Node(
            id=node_id,
            type="reservoir" if is_res else "junction",
            easting=coord[0] if coord else None,
            northing=coord[1] if coord else None,
            elevation_m=elevation[node_id],
            pressao_disp_mca=pressao_disp[node_id],
            pressao_est_mca=pressao_est[node_id],
            nivel_piezo_m=nivel_piezo[node_id],
            n_contrib=nc,
            base_demand_ls=demand,
            total_head_m=reservoir_head.get(node_id) if is_res else None,
        ))

    # 6. Montar Network ─────────────────────────────────────────────────────────
    network = Network(metadata=metadata, nodes=nodes, pipes=pipes)
    network.build_indexes()
    return network


def _prompt_missing_coordinates(network: Network, csv_path: Path) -> bool:
    """
    Para cada nó sem coordenadas, solicita E/N ao usuário interativamente
    e atualiza o nó. Retorna True se alguma coordenada foi inserida.

    Ao fim, oferece salvar as novas entradas no CSV original.
    """
    missing = [n for n in network.nodes if not n.has_coordinates()]
    if not missing:
        return False

    print(f"\nATENÇÃO: {len(missing)} nó(s) sem coordenadas no CSV.")
    print("Informe as coordenadas (sistema métrico projetado, ex: UTM).\n")

    new_entries: list[tuple[str, float, float]] = []

    for node in missing:
        print(f"  Nó '{node.id}':")
        while True:
            try:
                e = input("    Easting  (E): ").strip().replace(",", ".")
                n = input("    Northing (N): ").strip().replace(",", ".")
                node.easting  = float(e)
                node.northing = float(n)
                new_entries.append((node.id, node.easting, node.northing))
                break
            except ValueError:
                print("    Valor inválido — digite apenas números.\n")

    # Oferecer salvar no CSV
    resp = input(f"\nSalvar {len(new_entries)} nova(s) entrada(s) em '{csv_path}'? [s/N] ").strip().lower()
    if resp == "s":
        # Detecta delimitador existente
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
                delim = dialect.delimiter
            except csv.Error:
                delim = ","
        with open(csv_path, "a", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=delim)
            for node_id, east, north in new_entries:
                writer.writerow([node_id, east, north])
        print(f"  Salvo em '{csv_path}'.")

    return True


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Lê CSV + XLSX CORSAN e constrói objeto Network."
    )
    parser.add_argument(
        "config",
        type=Path,
        help="Caminho para o arquivo YAML de configuração do projeto",
    )
    args = parser.parse_args()

    config   = load_config(args.config)
    network  = build_network(config)

    csv_path = Path.cwd() / config["csv_path"]
    _prompt_missing_coordinates(network, csv_path)

    issues = network.validate()
    if issues:
        print("\nProblemas encontrados:", file=sys.stderr)
        for issue in issues:
            print(f"  • {issue}", file=sys.stderr)

    print(network.summary())

    output_json = config.get("output_json")
    if output_json:
        out_path = Path.cwd() / output_json
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(network.to_dict(), f, ensure_ascii=False, indent=2)
        print(f"\nJSON salvo em '{out_path}'")


if __name__ == "__main__":
    main()
